# -*- coding: utf-8 -*-
"""
Created on Mon May 26 09:46:53 2025

@author: Admin
"""

# 1. FILL NA 
import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook
import shutil 

input_folder = 'input_10_new'

for filename in os.listdir(input_folder):
    if not filename.lower().endswith('.xlsx'):
        continue

    filepath = os.path.join(input_folder, filename)
    print(f"Đang xử lý file: {filepath}")

    # Mở workbook, chọn sheet đầu tiên
    wb = load_workbook(filepath)
    ws = wb.active

    # --- BƯỚC 1: Đọc ô A1, tìm đoạn nằm giữa dấu nháy kép ---
    header_value = ws['A1'].value
    if not header_value:
        print("- Chưa có nội dung ở A1, bỏ qua file này.")
        continue

    # Tìm đoạn giữa hai dấu nháy kép (ví dụ: "01.K6.1.1.1 Sử dụng…")
    match = re.search(r'"([^"]+)"', header_value)
    if not match:
        print("- Không tìm thấy chuỗi trong dấu nháy kép ở A1, bỏ qua file này.")
        continue

    raw_skill = match.group(1).strip()   # Ví dụ: 01.K6.1.1.1Sử dụng được thuật ngữ...

    m2 = re.match(r'^([0-9A-Za-z\.]+)\s*(.+)$', raw_skill)
    if m2:
        code_part = m2.group(1)
        desc_part = m2.group(2)
    else:
        # Nếu không khớp, coi toàn bộ raw_skill là code, không có description
        code_part = raw_skill
        desc_part = ''    

    # Ghép lại thành "{code} - {description}" (nếu desc_part không rỗng)
    if desc_part:
        final_skill = f"{code_part} - {desc_part}"
    else:
        final_skill = code_part

    # --- BƯỚC 3: Xác định cột "Kĩ năng" ---
    # Giả sử dòng tiêu đề là dòng 4, tìm ô có chứa chữ 'kĩ năng' (không phân biệt hoa thường)
    header_row = 3
    skill_col = None
    for cell in ws[header_row]:
        if cell.value and 'kĩ năng' in str(cell.value).lower():
            skill_col = cell.column_letter
            break
    if not skill_col:
        # Nếu không tìm thấy tiêu đề "Kĩ năng", mặc định dùng cột H 
        skill_col = 'H'

    # --- BƯỚC 4: GHI ĐÈ toàn bộ các ô trong cột Kĩ năng ---
    # Từ hàng header_row+1 (thường là 5) cho đến ws.max_row, gán final_skill
    for row in range(header_row + 1, ws.max_row + 1):
        ws[f"{skill_col}{row}"].value = final_skill

    wb.save(filepath)
    print(f"- Đã cập nhật cột 'Kĩ năng' thành [ {final_skill} ] và lưu file: {filename}")

print("Hoàn tất xử lý toàn bộ file.")


# 2. RENAME INPUT FILE BY KC 

for filename in os.listdir(input_folder):
    if not filename.lower().endswith('.xlsx'):
        continue

    filepath = os.path.join(input_folder, filename)
    print(f"Đang xử lý tệp: {filepath}")

    try:
        # Mở workbook và chọn sheet đầu tiên
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        header_row = 3 
        skill_col = None

        # Tìm cột có tiêu đề chứa "kĩ năng"
        for cell in ws[header_row]:
            if cell.value and 'kĩ năng' in str(cell.value).lower():
                skill_col = cell.column_letter
                break

        if not skill_col:
            print(" - Không tìm thấy cột 'Kĩ năng', bỏ qua tệp này.")
            continue

        # Lấy giá trị từ ô đầu tiên trong cột 'Kĩ năng' (hàng 5)
        skill_cell = f"{skill_col}{header_row + 1}"
        skill_value = ws[skill_cell].value

        if not skill_value:
            print("- Ô 'Kĩ năng' đầu tiên trống, bỏ qua tệp này.")
            continue

        # Làm sạch giá trị để tạo tên tệp hợp lệ
        # Loại bỏ các ký tự không hợp lệ trong tên tệp
        cleaned_skill = re.sub(r'[\\/*?:"<>|]', '', str(skill_value)).strip()

        if not cleaned_skill:
            print("- Giá trị 'Kĩ năng' sau khi làm sạch trống, bỏ qua tệp này.")
            continue

        new_filename = f"{cleaned_skill}.xlsx"
        new_filepath = os.path.join(input_folder, new_filename)

        if os.path.exists(new_filepath):
            print(f"- Tệp '{new_filename}' đã tồn tại, bỏ qua để tránh ghi đè.")
            continue

        os.rename(filepath, new_filepath)
        print(f"- Đã đổi tên tệp thành: {new_filename}")

    except Exception as e:
        print(f"- Lỗi khi xử lý tệp '{filename}': {e}")

print("Hoàn tất xử lý tất cả các tệp.")

# 3. CONCAT 
input_folder = r'C:/Users/Admin/VDI/code/question_analysis/input_10_max_15'

excel_files = sorted(glob.glob(os.path.join(input_folder, '*.xlsx')))

df_list = []

for i, file in enumerate(excel_files):
    if i == 0:
        df0 = pd.read_excel(file, header=2)
        df_list.append(df0)
        # Lưu lại header để gán cho các file sau
        cols = df0.columns
    else:
        # Các file sau: bỏ 3 dòng đầu (2 dòng chung + 1 dòng header),
        # đọc phần dữ liệu mà không lấy header
        df = pd.read_excel(file, header=None, skiprows=3)
        # Gán lại header giống file đầu
        df.columns = cols
        df_list.append(df)

df = pd.concat(df_list, ignore_index=True)

def map_ten_ban_giao(row):
    qtype = row['Loại câu hỏi']
    has_img = row['Có sử dụng ảnh']
    content = str(row['Nội dung câu hỏi'])
    
    if qtype in ('multiple_choice', 'multiple_response_mmc'):
        if has_img == 0:
            return "Trắc nghiệm một/nhiều lựa chọn (dạng text)"
        else:
            return "Trắc nghiệm một/nhiều lựa chọn (dạng media)"
    
    if qtype == 'inline':
        # [đau,ốm] => dấu ngoặc vuông + dấu phẩy
        if re.search(r"\[[^\]]+|[^\]]+\]", content):
            return "Điền từ có gợi ý (lựa chọn)"
        if re.search(r"__[^_]+(?:/[^_]+)?__", content):
            return "Trả lời ngắn"
  
    other_map = {
        'likert': "Đúng sai dạng bảng",
        'multiple_drag_drop_matching': "Kéo thả",
        'drag_drop_matching': "Kéo thả",
        'matching_pair': "Ghép nối",
        'reorder_sentence_or_words': "Kéo thả" 
    }
    if qtype in other_map:
        return other_map[qtype]
    
    return qtype

df['Tên bàn giao'] = df.apply(map_ten_ban_giao, axis=1)

unmapped_df = df[df['Tên bàn giao'] == df['Loại câu hỏi']]

unmapped_types = unmapped_df['Loại câu hỏi'].unique()

print("Các loại câu hỏi chưa được map:")
for t in unmapped_types:
    print("-", t)
    
df = df.rename(columns={
    'Độ khó': 'Mức',
    'Tên bàn giao': 'Loại câu hỏi 2',
    'Có sử dụng ảnh': 'Ảnh',
    'Có sử dụng âm thanh': 'Âm thanh'
})

# 4. STATISTIC 

# Pivot đếm số lượng từng Mức (chuỗi) theo Kỹ năng
pivot_muc = pd.crosstab(df['Kĩ năng'], df['Mức'])\
              .reindex(columns=['Mức 1', 'Mức 2', 'Mức 3'], fill_value=0)

# Pivot đếm số lượng từng Loại câu hỏi (giá trị duy nhất trong 'Loại câu hỏi')
pivot_loai = pd.crosstab(df['Kĩ năng'], df['Loại câu hỏi 2'])

# Tổng hợp ảnh & âm thanh (giả sử cột 0/1 hoặc True/False)
sum_av = df.groupby('Kĩ năng')[['Ảnh', 'Âm thanh']].sum()

# Đếm số thẻ (loại bỏ NaN)
count_tags = (df.dropna(subset=['Thẻ'])
                .groupby('Kĩ năng')['Thẻ']
                .count()
                .rename('Số thẻ'))

# Ghép tất cả lại
result = (
    pivot_muc
    .join(pivot_loai, how='outer')
    .join(sum_av, how='outer')
    .join(count_tags, how='outer')
    .fillna(0)
    .astype(int)
    .reset_index()
)

result = result.rename(columns={
    'Ảnh': 'Tổng ảnh',
    'Âm thanh': 'Tổng âm thanh'
})

os.makedirs('output_6', exist_ok=True) 

# Lưu df vào CSV, không lưu index, mã hóa utf-8 với BOM (nếu cần Excel đọc ngon hơn)
result.to_csv('output_6/statistics_2.csv', index=False, encoding='utf-8-sig')
df.to_csv('output_6/check_nan.csv', index=False, encoding='utf-8-sig')

df['Kĩ năng'].isnull().sum()
df['Kĩ năng'].value_counts() 

# 5. GET MAX SKILLS 

top_skills = (df['Kĩ năng'].value_counts().nlargest(20).index.tolist())

input_folder = 'input_10'
output_folder = 'input_10_max_20'

os.makedirs(output_folder, exist_ok=True)

# Duyệt qua các kỹ năng và sao chép file 
for skill in top_skills:
    filename = f"{skill}.xlsx"
    src_path = os.path.join(input_folder, filename)
    if os.path.isfile(src_path):
        shutil.copy2(src_path, output_folder)
        print(f"[OK] Đã copy: {filename}")
    else:
        print(f"[WARN] Không tìm thấy file: {filename}")

# 6. CREATE NEW DF & CSV 

df = df[['Nội dung câu hỏi', 'Kĩ năng']] 
df.rename(columns = {'Nội dung câu hỏi': 'Content', 'Kĩ năng': 'KC'}, inplace=True)
df.to_csv('output_10/nbc_data_top_15.csv', index=False, encoding='utf-8-sig')





